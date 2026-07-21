"""Fetch the ONS "General government annual expenditure: ESA Table 11"
dataset (COFOG-classified spending by function) and write
data-uk-spending-cofog.json: the most recent year's spending broken down
by category (health, education, social protection, defence, etc.), in £m
and as % of total.

Run:  python3 fetch_uk_spending_cofog.py
No API key needed -- public ONS download, no auth.

SOURCE: https://www.ons.gov.uk/economy/governmentpublicsectorandtaxes/
publicspending/datasets/esatable11annualexpenditureofgeneralgovernment/current
This is confirmed to have a STABLE download URL (unlike INAC01, whose
filename changes every release) -- confirmed live 2026-07-21, the
"current" edition sits at a fixed filename
(esatable11generalgovernment.xlsx) that ONS updates in place, so this
script hits that URL directly rather than scraping the landing page for
a link first. Confirmed OGL v3.0 licensed (stated on the dataset page).

FREQUENCY WARNING, worth surfacing on the frontend: this is ANNUAL data
with a real lag -- the edition live as of 2026-07-21 was released
23 April 2026 but (per ONS's own FOI page on this topic) the underlying
outturn data itself typically runs ~2 years behind the publication date.
This is a genuinely different freshness profile than the rest of the
site and should be labelled as such, not presented as if it were as
current as the monthly/quarterly series everywhere else.

NOT YET CONFIRMED: the internal layout of the spreadsheet (which
row/column holds which COFOG category, how many years of columns exist).
Parsing below searches for the ten standard COFOG level-1 category names
(these are internationally standardised, not a guess) rather than
hardcoding row/column positions, and dumps raw content for diagnosis if
matching fails -- same approach used successfully for the BoE and INAC01
spreadsheets this session.
"""

from __future__ import annotations

import io
import json
import sys
from datetime import datetime, timezone

import requests

CURRENT_XLSX_URL = ("https://www.ons.gov.uk/file?uri=/economy/governmentpublicsectorandtaxes/"
                     "publicspending/datasets/esatable11annualexpenditureofgeneralgovernment/"
                     "current/esatable11generalgovernment.xlsx")

# The ten standard COFOG level-1 divisions -- internationally standardised
# by the UN Statistics Division, not something this site is guessing at.
COFOG_LABELS = {
    "general_public_services": ["general public services"],
    "defence": ["defence"],
    "public_order_safety": ["public order and safety"],
    "economic_affairs": ["economic affairs"],
    "environmental_protection": ["environmental protection"],
    "housing_community": ["housing and community amenities"],
    "health": ["health"],
    "recreation_culture": ["recreation, culture and religion", "recreation culture and religion"],
    "education": ["education"],
    "social_protection": ["social protection"],
}


def _grid_from_openpyxl(ws):
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def _parse_grid(grid: list[list]) -> dict | None:
    found = {}
    for r, row in enumerate(grid):
        for c, val in enumerate(row):
            if not isinstance(val, str):
                continue
            label = val.strip().lower()
            for key, patterns in COFOG_LABELS.items():
                if key in found:
                    continue
                if any(p in label for p in patterns):
                    last_val = None
                    for cc in range(c, len(row)):
                        v = row[cc]
                        if isinstance(v, (int, float)):
                            last_val = v
                    if last_val is not None:
                        found[key] = last_val
                        print(f"  [cofog] matched '{key}' at row {r+1} -> {last_val}")

    if len(found) < 6:
        print(f"  [cofog] only matched {len(found)}/10 categories via label search — "
              f"dumping first 40 rows, cols 1-14 for diagnosis:")
        for r in range(min(40, len(grid))):
            vals = [repr(v) for v in grid[r][:14]]
            print(f"  [cofog] row {r+1}: {vals}")
        return None

    total = sum(found.values())
    if total <= 0:
        print("  [cofog] matched categories but values summed to zero")
        return None

    categories = [{"key": k, "value_gbp_m": v, "share_pct": round(v / total * 100, 1)}
                  for k, v in found.items()]
    categories.sort(key=lambda c: c["value_gbp_m"], reverse=True)
    print(f"  [cofog] {len(categories)} categories, total={total}")
    return {"categories": categories, "total_gbp_m": total}


def fetch_and_parse() -> dict | None:
    try:
        r = requests.get(CURRENT_XLSX_URL, timeout=60, headers={"User-Agent": "economic-atlas/0.1"})
        print(f"  [cofog] status={r.status_code}")
        r.raise_for_status()
    except Exception as exc:
        print(f"  [cofog] request failed: {exc}")
        return None

    try:
        import openpyxl
    except ImportError:
        print("  [cofog] openpyxl not installed — skipping")
        return None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    except Exception as exc:
        print(f"  [cofog] couldn't open workbook: {exc}")
        return None

    print(f"  [cofog] sheets: {wb.sheetnames}")
    # Try every sheet, not just the first -- COFOG workbooks sometimes
    # split central/local/general government across separate tabs, and
    # this dataset is specifically the "general government" edition, which
    # may not necessarily be the first sheet.
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  [cofog] trying sheet '{sheet_name}': {ws.max_row} rows x {ws.max_column} cols")
        grid = _grid_from_openpyxl(ws)
        result = _parse_grid(grid)
        if result:
            result["sheet"] = sheet_name
            return result
    print("  [cofog] no sheet matched enough COFOG category labels")
    return None


def main() -> int:
    result = fetch_and_parse()
    if not result:
        print("FAIL  cofog  no usable data — leaving any previously-fetched file in place")
        return 0

    out = {
        "updated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": "ONS, General government annual expenditure: ESA Table 11 (COFOG)",
        "frequency_note": "Annual, with a real lag behind the publication date -- not comparable "
                           "in freshness to the site's monthly/quarterly series.",
        **result,
    }
    with open("data-uk-spending-cofog.json", "w") as f:
        json.dump(out, f)
    print(f"  ok  cofog  wrote {len(result['categories'])} categories")
    return 0


if __name__ == "__main__":
    sys.exit(main())
