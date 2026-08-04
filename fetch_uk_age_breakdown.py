"""Fetch the ONS A05 SA dataset (employment, unemployment and economic
inactivity by age group, seasonally adjusted) and write
data-uk-age-breakdown.json: the most recent period's unemployment and
inactivity rates broken down by age band.

Run:  python3 fetch_uk_age_breakdown.py
No API key needed -- public ONS download, no auth.

SOURCE: https://www.ons.gov.uk/employmentandlabourmarket/peopleinwork/
employmentandemployeetypes/datasets/employmentunemploymentandeconomic
inactivitybyagegroupseasonallyadjusteda05sa/current
Confirmed OGL v3.0. Same mechanism as INAC01: filename changes every
release (confirmed live 2026-07-21, current was "a05saapr2026.xls"), so
this scrapes the landing page for the current edition link rather than
hardcoding a filename.

NOT YET CONFIRMED: the internal sheet/column layout. This file is ~1.5MB,
likely covers many age bands (16-17, 18-24, 25-34, 35-49, 50-64, 65+, and
various combined bands) across levels and rates for employment,
unemployment and inactivity -- almost certainly a larger, multi-sheet
workbook than INAC01. Unlike INAC01/COFOG, no dataset identifier codes
have been confirmed for this file yet, so this tries two strategies in
order: (1) search for a single cell combining a metric name with an
age-band label, same as before; (2) if that finds nothing, look for a
row of short alphanumeric codes (e.g. LF64-style), the same "Dataset
identifier code" row convention CONFIRMED working on the sibling INAC01
script from the same ONS team -- this doesn't assume any specific codes,
it just surfaces candidate code rows explicitly in the log rather than
leaving them buried in a blind row dump. If neither strategy finds
anything, this still dumps raw content for diagnosis, same discipline as
every other fetch script on this site: search for anchors, don't guess
positions, and show real failures rather than guessing blindly.

IMPORTANT: the age bands actually available in this file may not match
"25-49"/"50-64" (the illustrative placeholders currently used on the
site) -- ONS's actual published bands for A05 might be split differently
(e.g. 18-24, 25-34, 35-49, 50-64, 65+). Whatever bands are actually found
will be written out with their real labels; the frontend should be
updated to match once this is confirmed live, not the other way around.
"""

from __future__ import annotations

import io
import json
import re
import sys
from datetime import datetime, timezone

import requests

DATASET_PAGE = ("https://www.ons.gov.uk/employmentandlabourmarket/peopleinwork/"
                 "employmentandemployeetypes/datasets/"
                 "employmentunemploymentandeconomicinactivitybyagegroupseasonallyadjusteda05sa/current")

# Age bands we're hoping to find -- matched flexibly against header text
# since no identifier codes are confirmed yet for this file.
AGE_BAND_PATTERNS = ["16-17", "18-24", "16-24", "25-34", "35-49", "25-49", "50-64", "65+"]
METRIC_PATTERNS = {
    "unemployment_rate": ["unemployment rate"],
    "inactivity_rate": ["inactivity rate", "economic inactivity rate"],
}


def _grid_from_openpyxl(ws):
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def _grid_from_xlrd(sheet):
    return [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]


def _to_float(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s in ("", "..", "-", ":", "n/a"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def find_current_xls_link(html: str) -> str | None:
    m = re.search(r'href="(/file\?uri=[^"]+\.xls)"', html)
    if m:
        return "https://www.ons.gov.uk" + m.group(1).replace("&amp;", "&")
    return None


def _dump(grid, label, n=40, cols=14):
    print(f"  [a05] {label} — dumping first {n} rows, cols 1-{cols}:")
    for r in range(min(n, len(grid))):
        print(f"  [a05] row {r+1}: {[repr(v) for v in grid[r][:cols]]}")


def fetch_and_parse() -> dict | None:
    try:
        r = requests.get(DATASET_PAGE, timeout=60, headers={"User-Agent": "economic-atlas/0.1"})
        print(f"  [a05] page status={r.status_code}")
        r.raise_for_status()
    except Exception as exc:
        print(f"  [a05] page request failed: {exc}")
        return None

    xls_url = find_current_xls_link(r.text)
    if not xls_url:
        print("  [a05] couldn't find a current-edition .xls link on the dataset page")
        return None
    print(f"  [a05] current edition: {xls_url}")

    try:
        xls_r = requests.get(xls_url, timeout=60, headers={"User-Agent": "economic-atlas/0.1"})
        print(f"  [a05] xls status={xls_r.status_code}")
        xls_r.raise_for_status()
    except Exception as exc:
        print(f"  [a05] xls download failed: {exc}")
        return None

    sheet_order = []
    get_grid = None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xls_r.content), data_only=True)
        print(f"  [a05] opened via openpyxl, sheets: {wb.sheetnames}")
        sheet_order = list(wb.sheetnames)
        get_grid = lambda name: _grid_from_openpyxl(wb[name])
    except Exception as exc:
        print(f"  [a05] openpyxl couldn't open it ({exc}) — trying xlrd "
              f"(likely genuine legacy .xls format)")
        try:
            import xlrd
            book = xlrd.open_workbook(file_contents=xls_r.content)
            print(f"  [a05] opened via xlrd, sheets: {book.sheet_names()}")
            sheet_order = list(book.sheet_names())
            get_grid = lambda name: _grid_from_xlrd(book.sheet_by_name(name))
        except ImportError:
            print("  [a05] xlrd not installed either")
            return None
        except Exception as exc2:
            print(f"  [a05] xlrd also failed to open it: {exc2}")
            return None

    # We genuinely don't know which sheet(s) matter yet -- print the full
    # sheet list so the first live run tells us what's actually in here,
    # then try each one.
    last_grid = None
    last_sheet = None
    for sheet_name in sheet_order:
        grid = get_grid(sheet_name)
        last_grid, last_sheet = grid, sheet_name
        print(f"  [a05] trying sheet '{sheet_name}': {len(grid)} rows")

        # Strategy 1: look for header cells combining a metric with an
        # age band in the SAME cell -- collect candidate (row, col,
        # metric, band) hits across the whole sheet, since we don't know
        # the layout yet.
        hits = []
        for r, row in enumerate(grid[:60]):  # header block is almost certainly near the top
            for c, val in enumerate(row):
                if not isinstance(val, str):
                    continue
                label = val.strip().lower()
                for metric, mpatterns in METRIC_PATTERNS.items():
                    if any(mp in label for mp in mpatterns):
                        for band in AGE_BAND_PATTERNS:
                            if band.lower() in label:
                                hits.append((r, c, metric, band))
        if hits:
            print(f"  [a05] sheet '{sheet_name}': found {len(hits)} metric+band header hits: {hits[:20]}")
            # Not yet confirmed enough to auto-extract values -- surface
            # what was found so the next round can wire up extraction
            # against the real confirmed positions, rather than guessing
            # at a value-extraction rule with no confirmed layout.
            continue

        # Strategy 2: the sibling INAC01 script (same ONS team, same
        # Labour Force Survey family, CONFIRMED working) uses a distinct
        # "Dataset identifier code" row instead of free-text headers --
        # a row of short alphanumeric codes (e.g. LF64, LF66) that map
        # columns to categories far less ambiguously than label text.
        # Strategy 1 alone has found zero hits across every run so far,
        # which is exactly what you'd expect if this file uses the same
        # identifier-code convention instead of (or as well as) text
        # labels. This doesn't assume any specific codes -- A05 SA's own
        # codes haven't been confirmed yet -- it just surfaces any
        # candidate code row explicitly, rather than leaving it buried
        # in a blind dump of the first 40 rows.
        code_pattern = re.compile(r"^[A-Z][A-Z0-9]{2,4}$")
        code_hits = []
        for r, row in enumerate(grid[:60]):
            row_codes = [(c, val.strip()) for c, val in enumerate(row)
                         if isinstance(val, str) and code_pattern.match(val.strip())]
            if len(row_codes) >= 3:
                code_hits.append((r, row_codes))
        if code_hits:
            print(f"  [a05] sheet '{sheet_name}': no metric+band text hits, but found "
                  f"{len(code_hits)} row(s) that look like an ONS dataset-identifier-code "
                  f"row (same pattern as the working INAC01 script) -- these are the codes "
                  f"to map next, not yet confirmed against ONS's own code list:")
            for r, row_codes in code_hits:
                print(f"  [a05] row {r+1} candidate codes: {row_codes}")
            # The codes alone are useless without whatever's labelling them --
            # ONS timeseries workbooks almost always carry the human-readable
            # series title in a row just above (or occasionally below) the
            # CDID code row. Dumping the codes and moving on (as this used to
            # do) threw away exactly the information needed to map them.
            # Surface the rows immediately around the first code row too, so
            # the *next* run gives us the title text directly instead of
            # requiring a third round-trip.
            first_code_row = code_hits[0][0]
            dump_start = max(0, first_code_row - 6)
            dump_end = min(len(grid), first_code_row + 4)
            print(f"  [a05] sheet '{sheet_name}': dumping rows {dump_start + 1}-{dump_end} "
                  f"around the code row for title context:")
            for r in range(dump_start, dump_end):
                print(f"  [a05] row {r+1}: {[repr(v) for v in grid[r][:20]]}")
            continue

    print("  [a05] no sheet produced confirmed metric+age-band header hits — "
          "dumping the last sheet tried (not sheet_order[0], which was often an "
          "unrelated 'Note'/disclaimer sheet and gave nothing useful):")
    if last_grid is not None:
        _dump(last_grid, f"sheet '{last_sheet}'")
    return None


def main() -> int:
    result = fetch_and_parse()
    if not result:
        print("FAIL  a05  no usable data yet — leaving any previously-fetched file in place. "
              "This is expected on a first run against a real, unconfirmed file layout; "
              "check the sheet names and any header hits printed above.")
        return 0

    out = {
        "updated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": "ONS A05 SA: Employment, unemployment and economic inactivity by age group (seasonally adjusted)",
        **result,
    }
    with open("data-uk-age-breakdown.json", "w") as f:
        json.dump(out, f)
    print(f"  ok  a05  wrote age breakdown")
    return 0


if __name__ == "__main__":
    sys.exit(main())
