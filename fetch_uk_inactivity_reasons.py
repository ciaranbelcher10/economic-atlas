"""Fetch the ONS INAC01 SA dataset (economic inactivity by reason,
seasonally adjusted) and write data-uk-inactivity-reasons.json: the most
recent period's breakdown of why people are economically inactive
(long-term/temporary sick, student, looking after family, retired,
other).

Run:  python3 fetch_uk_inactivity_reasons.py
No API key needed -- public ONS download, no auth.

SOURCE: https://www.ons.gov.uk/employmentandlabourmarket/peoplenotinwork/
economicinactivity/datasets/economicinactivitybyreasonseasonallyadjustedinac01sa
Filename changes every release, so this scrapes the landing page for the
current "current edition" .xls link rather than hardcoding one.
CONFIRMED OGL v3.0 licensed (stated on the dataset page).

CONFIRMED live (2026-07-21) real layout, via two rounds of diagnostic
dumps -- this is NOT a one-row-per-category table like the COFOG
workbook. It's:
  - Legacy binary .xls (openpyxl can't open it; xlrd can -- confirmed,
    not assumed)
  - Four sheets: 'Note' (a disclaimer, not data), 'People' (all persons --
    the one we want), 'Men', 'Women'
  - Standard ONS time-series layout: one row per period ("Mar-May 1992",
    "Apr-Jun 1992", ...), one column per category, values in thousands
  - A "Dataset identifier code" row gives each category column a stable
    ONS series code -- these are used as the anchor instead of label
    text, since they're less ambiguous:
      LF64 = Student
      LF66 = Looking after family / home
      LF68 = Temp sick
      LF6A = Long-term sick
      LFM3 = Discouraged workers
      LF6C = Retired
      LF6E = Other
  - Early rows use ".." as a placeholder for "not collected yet" --
    parsing must skip those and find the actual most recent row with
    real numbers, not just the literal last row.
  - The published "long-term/temporary sickness" figure quoted in ONS's
    own bulletins (e.g. "33% of inactive people") is Temp sick + Long-term
    sick added together, not a single native column -- so this script
    sums LF68 and LF6A rather than treating them as separate categories.
    "Discouraged workers" (LFM3) is folded into "Other" to keep to the
    same five-category framing already used elsewhere on the site.
"""

from __future__ import annotations

import io
import json
import re
import sys
from datetime import datetime, timezone

import requests

DATASET_PAGE = ("https://www.ons.gov.uk/employmentandlabourmarket/peoplenotinwork/"
                 "economicinactivity/datasets/economicinactivitybyreasonseasonallyadjustedinac01sa")

# ONS dataset identifier codes -> our category keys. Confirmed live
# (2026-07-21) against the actual "Dataset identifier code" row in the
# real workbook, not guessed.
CODE_TO_CATEGORY = {
    "LF64": "student",
    "LF66": "looking_after_family",
    "LF6A": "long_term_sick",   # summed with LF68 below
    "LF68": "long_term_sick",
    "LFM3": "other",            # "discouraged workers" folded into other
    "LF6C": "retired",
    "LF6E": "other",
}


def _grid_from_openpyxl(ws):
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def _grid_from_xlrd(sheet):
    return [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]


def _to_float(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s in ("", "..", "-", ":", "n/a"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def _parse_timeseries_grid(grid, quiet=False):
    """Parse the real INAC01 layout: a 'Dataset identifier code' row maps
    columns to ONS series codes; data rows below give one period per row,
    with '..' placeholders for periods not yet collected for a given
    column. Finds the most recent row where every needed column has a
    real number, working backward from the bottom of the sheet."""
    code_row_idx = None
    col_for_code = {}
    for r, row in enumerate(grid):
        for c, val in enumerate(row):
            if isinstance(val, str) and val.strip() in CODE_TO_CATEGORY:
                col_for_code[val.strip()] = c
                code_row_idx = r
    if not col_for_code or len(col_for_code) < 4:
        if not quiet:
            print(f"  [inac01] found identifier-code row but only matched "
                  f"{len(col_for_code)}/6 codes — dumping rows around it for diagnosis:")
            start = max(0, (code_row_idx or 0) - 2)
            for r in range(start, min(start + 15, len(grid))):
                print(f"  [inac01] row {r+1}: {[repr(v) for v in grid[r][:12]]}")
        return None

    period_col = 0  # period label is always the first column in this layout
    needed_cols = list(col_for_code.values())

    # Search from the bottom up for the most recent row where every needed
    # column has a real (non-placeholder) number.
    for r in range(len(grid) - 1, code_row_idx, -1):
        row = grid[r]
        if len(row) <= max(needed_cols):
            continue
        values = {code: _to_float(row[col]) for code, col in col_for_code.items()}
        if all(v is not None for v in values.values()):
            period = row[period_col]
            totals = {}
            for code, val in values.items():
                cat = CODE_TO_CATEGORY[code]
                totals[cat] = totals.get(cat, 0.0) + val
            total = sum(totals.values())
            if total <= 0:
                continue
            print(f"  [inac01] using period '{period}' (row {r+1}): {totals}")
            reasons = [{"key": k, "value": v, "share_pct": round(v / total * 100, 1)}
                       for k, v in totals.items()]
            return {"reasons": reasons, "total": total, "period": period}

    if not quiet:
        print("  [inac01] identifier codes matched but no row had complete real data — "
              "dumping the last 15 rows for diagnosis:")
        for r in range(max(0, len(grid) - 15), len(grid)):
            print(f"  [inac01] row {r+1}: {[repr(v) for v in grid[r][:12]]}")
    return None


def find_current_xls_link(html):
    m = re.search(r'href="(/file\?uri=[^"]+\.xls)"', html)
    if m:
        return "https://www.ons.gov.uk" + m.group(1).replace("&amp;", "&")
    return None


def fetch_and_parse():
    try:
        r = requests.get(DATASET_PAGE, timeout=60, headers={"User-Agent": "economic-atlas/0.1"})
        print(f"  [inac01] page status={r.status_code}")
        r.raise_for_status()
    except Exception as exc:
        print(f"  [inac01] page request failed: {exc}")
        return None

    xls_url = find_current_xls_link(r.text)
    if not xls_url:
        print("  [inac01] couldn't find a current-edition .xls link on the dataset page")
        return None
    print(f"  [inac01] current edition: {xls_url}")

    try:
        xls_r = requests.get(xls_url, timeout=60, headers={"User-Agent": "economic-atlas/0.1"})
        print(f"  [inac01] xls status={xls_r.status_code}")
        xls_r.raise_for_status()
    except Exception as exc:
        print(f"  [inac01] xls download failed: {exc}")
        return None

    sheet_order = []
    get_grid = None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xls_r.content), data_only=True)
        print(f"  [inac01] opened via openpyxl, sheets: {wb.sheetnames}")
        sheet_order = list(wb.sheetnames)
        get_grid = lambda name: _grid_from_openpyxl(wb[name])
    except Exception as exc:
        print(f"  [inac01] openpyxl couldn't open it ({exc}) — trying xlrd "
              f"(likely genuine legacy .xls format)")
        try:
            import xlrd
            book = xlrd.open_workbook(file_contents=xls_r.content)
            print(f"  [inac01] opened via xlrd, sheets: {book.sheet_names()}")
            sheet_order = list(book.sheet_names())
            get_grid = lambda name: _grid_from_xlrd(book.sheet_by_name(name))
        except ImportError:
            print("  [inac01] xlrd not installed either")
            return None
        except Exception as exc2:
            print(f"  [inac01] xlrd also failed to open it: {exc2}")
            return None

    # Only 'People' (all-persons) is acceptable here -- this dataset is
    # published as 'all persons' on the site, and Men/Women are a
    # different statistic, not a same-population fallback. Silently
    # substituting one sheet for the other produced a real live bug:
    # a 'People' parse failure fell through to 'Women' and the result
    # was still written out labelled "all persons". If 'People' can't be
    # parsed, that's a genuine failure -- fail loudly (main() leaves the
    # previously-fetched file in place) rather than mislabel gendered
    # data as the whole population, matching the stricter pattern
    # fetch_uk_age_breakdown.py already uses for the same reason.
    if "People" not in sheet_order:
        print(f"  [inac01] no 'People' sheet found (have: {sheet_order}) -- layout has changed")
        return None

    grid = get_grid("People")
    print(f"  [inac01] trying sheet 'People': {len(grid)} rows")
    result = _parse_timeseries_grid(grid, quiet=False)
    if result:
        result["sheet"] = "People"
        return result

    print("  [inac01] 'People' sheet found but produced no usable row.")
    return None


def main():
    result = fetch_and_parse()
    if not result:
        print("FAIL  inac01  no usable data — leaving any previously-fetched file in place")
        return 0

    out = {
        "updated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": "ONS INAC01 SA: Economic inactivity by reason (seasonally adjusted), all persons",
        **result,
    }
    with open("data-uk-inactivity-reasons.json", "w") as f:
        json.dump(out, f)
    print(f"  ok  inac01  wrote {len(result['reasons'])} reason categories for {result.get('period')}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
