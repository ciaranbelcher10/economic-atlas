"""Fetch the Bank of England's official MPC voting-history spreadsheet and
write data-uk-mpc-votes.json: per-meeting vote breakdowns (how many members
backed the decision vs. each alternative), for the Inflation tab's MPC
vote-record module.

Run:  python3 fetch_boe_mpc.py
No API key needed -- this is a public file download, no auth.

SOURCE: https://www.bankofengland.co.uk/-/media/boe/files/monetary-policy-
summary-and-minutes/mpcvoting.xlsx -- the Bank's own official long-format
voting record (one row per meeting-member pair), confirmed to exist via a
documented R package wrapper (the "boe" CRAN package's boe_mpc_votes()
function, which describes exactly this file's columns: meeting date,
member name, the Bank Rate that member voted for, and the committee's
actual decision).

NOT YET CONFIRMED LIVE: this sandbox can't reach bankofengland.co.uk (not
in the allowed outbound domains here), so the exact column headers/layout
have never actually been inspected -- only inferred from that third-party
package's documented schema. GitHub Actions runners have normal internet
access and will actually hit the real file. Parsing below is deliberately
defensive (keyword-matches header names rather than assuming fixed
column positions/order) and logs exactly what it finds, so the first live
run's [boe-mpc] log lines will show definitively whether the assumed
layout was right -- adjust from there rather than guessing further blind.

WHY NO NAMED INDIVIDUAL VOTES IN THE OUTPUT (yet): the source data does
include each member's name, but before publishing real people's individual
votes, the parsed output should be spot-checked against the Bank's own
published minutes (e.g. https://www.bankofengland.co.uk/monetary-policy-
summary-and-minutes/...) at least once, given how much worse a
misattribution of a real, identifiable person's vote is than an
approximation elsewhere on the site. This first cut only emits AGGREGATE
counts per meeting (how many voted hold/hike/cut), which the aggregate
Monetary Policy Summary text also states independently -- giving a
built-in cross-check. Names can be added once that spot-check happens.
"""

from __future__ import annotations

import io
import json
import re
import sys
from datetime import datetime, timezone

import requests

VOTING_XLSX_URL = ("https://www.bankofengland.co.uk/-/media/boe/files/"
                    "monetary-policy-summary-and-minutes/mpcvoting.xlsx")

# How many most-recent meetings to keep in the output history.
HISTORY_LIMIT = 16


def _find_header_row(ws) -> tuple[int, dict[str, int]] | None:
    """Scan the first few rows for a header row, matching column purpose
    by keyword rather than assuming a fixed layout. Returns
    (header_row_index, {purpose: col_index}) or None if not found."""
    wanted = {
        "date": ("date", "meeting"),
        "member": ("name", "member"),
        "vote": ("vote", "voted"),
        "decision": ("decision", "outcome", "bank rate"),
    }
    for row_idx in range(1, 6):
        cols: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if not v:
                continue
            label = str(v).strip().lower()
            for purpose, keywords in wanted.items():
                if purpose in cols:
                    continue
                if any(kw in label for kw in keywords):
                    cols[purpose] = col_idx
        if "date" in cols and "vote" in cols:
            print(f"  [boe-mpc] header row {row_idx}: {cols}")
            return row_idx, cols
    return None


def fetch_and_parse() -> list[dict] | None:
    try:
        r = requests.get(VOTING_XLSX_URL, timeout=60,
                         headers={"User-Agent": "economic-atlas/0.1"})
        print(f"  [boe-mpc] status={r.status_code}")
        r.raise_for_status()
    except Exception as exc:
        print(f"  [boe-mpc] request failed: {exc}")
        return None

    try:
        import openpyxl
    except ImportError:
        print("  [boe-mpc] openpyxl not installed — skipping")
        return None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    except Exception as exc:
        print(f"  [boe-mpc] couldn't open as xlsx: {exc}")
        return None

    print(f"  [boe-mpc] sheets in workbook: {wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]
    print(f"  [boe-mpc] sheet '{ws.title}': {ws.max_row} rows x {ws.max_column} cols")

    found = _find_header_row(ws)
    if not found:
        # CONFIRMED live (2026-07-20): 3818 rows x 191 cols is nowhere near
        # the simple long-format (meeting, member, vote, decision) table
        # assumed from the R package docs -- that shape suggests a wide/
        # pivot layout instead (e.g. one column per meeting), or the real
        # data sits on a different sheet than the first one. Rather than
        # guess a third time, dump the actual raw corner of the sheet so
        # the real layout can be read directly from the next log rather
        # than inferred blind.
        print("  [boe-mpc] couldn't find a recognisable header row via keyword "
              "matching -- dumping raw content instead (rows 1-10, cols 1-15):")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_vals = []
            for col_idx in range(1, min(16, ws.max_column + 1)):
                v = ws.cell(row=row_idx, column=col_idx).value
                row_vals.append(repr(v) if v is not None else "")
            print(f"  [boe-mpc] row {row_idx}: {row_vals}")
        return None
    header_row, cols = found

    # meeting_date -> {"decision": float, "votes": [rate, rate, ...]}
    by_meeting: dict[str, dict] = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_date = ws.cell(row=row_idx, column=cols["date"]).value
        raw_vote = ws.cell(row=row_idx, column=cols["vote"]).value
        raw_decision = ws.cell(row=row_idx, column=cols.get("decision", cols["vote"])).value
        if raw_date is None or raw_vote is None:
            continue
        if hasattr(raw_date, "strftime"):
            date_str = raw_date.strftime("%Y-%m-%d")
        else:
            date_str = str(raw_date).strip()
            m = re.search(r"(\d{4})-(\d{2})-(\d{2})", date_str)
            if not m:
                continue
        try:
            vote_rate = float(raw_vote)
        except (TypeError, ValueError):
            continue
        try:
            decision_rate = float(raw_decision) if raw_decision is not None else vote_rate
        except (TypeError, ValueError):
            decision_rate = vote_rate

        entry = by_meeting.setdefault(date_str, {"decision": decision_rate, "votes": []})
        entry["votes"].append(vote_rate)
        entry["decision"] = decision_rate  # last non-null wins; should be constant per meeting

    if not by_meeting:
        print("  [boe-mpc] parsed header but found 0 usable data rows")
        return None

    meetings = []
    for date_str, entry in by_meeting.items():
        decision = entry["decision"]
        votes = entry["votes"]
        hold = sum(1 for v in votes if abs(v - decision) < 1e-9)
        hike = sum(1 for v in votes if v > decision + 1e-9)
        cut = sum(1 for v in votes if v < decision - 1e-9)
        meetings.append({
            "date": date_str, "decision_rate": decision,
            "hold": hold, "hike": hike, "cut": cut, "total_votes": len(votes),
        })
    meetings.sort(key=lambda m: m["date"], reverse=True)

    # Fill in "previous rate" for each meeting so the frontend can say
    # "held at X%" vs "cut from Y% to X%" without a second lookup.
    for i, m in enumerate(meetings):
        m["previous_rate"] = meetings[i + 1]["decision_rate"] if i + 1 < len(meetings) else None

    print(f"  [boe-mpc] {len(meetings)} meetings parsed, "
          f"latest {meetings[0]['date']} ({meetings[0]['hold']}-{meetings[0]['hike']}-{meetings[0]['cut']})")
    return meetings[:HISTORY_LIMIT]


def main() -> int:
    meetings = fetch_and_parse()
    if not meetings:
        print("FAIL  boe-mpc  no usable data — leaving any previously-fetched file in place")
        return 0

    out = {
        "updated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": "Bank of England, official MPC voting-history spreadsheet",
        "meetings": meetings,
    }
    with open("data-uk-mpc-votes.json", "w") as f:
        json.dump(out, f)
    print(f"  ok  boe-mpc  wrote {len(meetings)} meetings")
    return 0


if __name__ == "__main__":
    sys.exit(main())
